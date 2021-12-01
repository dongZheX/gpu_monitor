# !/usr/bin/env python
# -*-coding:utf-8 -*-
# @Time    : 2021/11/29 20:11
# @Author  : dongZheX
# @Version : python3.7
# @Desc    : $END$
import win32api
from openpyxl import load_workbook
import argparse
import time
from ExcelOp import ExcelOp
from SSH import SSHConnection
import json
import numpy as np
import paramiko
import os
from pyfiglet import Figlet
from gpu_thread import gpu_runner
from task import Task
column = ['id','dir','command',	'server','gpu','free','Priority','state','start_time','running_time',"ps_id","used"]
task_id = 0
os.system("cls")
servers = json.load(open("configs/server.json"))
servers_name = list(servers.keys())


def welcome():
    f = Figlet(width=2000)
    print(f.renderText("Welcome to GPU Monitor!"))


def print_menu():

    print("1- Nvidia-Monitor")
    print("2- Task Manager")
    print("3- Nvidia-Runner")
    print("4- exit")
    ia = input("Enter Value 1-4: ")
    try:
        ia = int(ia)
        if ia < 1 or ia > 4:
            print("Error Input,please Enter Value 1-4:", end='')
        else:
            return ia
    except ValueError:
        print("Error Input,please Enter Value 1-4:", end='')

    ia = input("")
    while True:
        try:
            ia = int(ia)
            if ia < 1 or ia > 4:
                ia = input("Error Input,please Enter Value 1-4: ")
            else:
                break
        except ValueError:
            ia = input("Error Input,please Enter Value 1-4: ")
    return ia

def monitor():

    command = "nvidia-smi --query-gpu=index,temperature.gpu,utilization.gpu,memory.used,memory.total,gpu_name --format=csv,nounits,noheader"
    conns = [SSHConnection(servers[name]['hostname'], port=servers[name]['port'], username=servers[name]['username']
                           , password=servers[name]['password']) for name in servers_name]
    while True:
        try:
            time.sleep(2)
            retss = []
            for i, conn in enumerate(conns):
                ret = str(conns[i].exec_command(command))
                start = ret.find('\'')
                end = ret.find('\'', start + 1)
                ret = ret[start + 1:end]
                ret = ret.replace(' ', '')
                ret = ret.replace('\\n', ',')[:-1]
                # print(ret)
                retss.append(np.asarray(ret.split(',')).reshape(-1, 6))
                # print(rets)
            os.system('cls')
            for i, rets in enumerate(retss):
                print("====================================%5s====================================" % servers_name[i])
                print("%--25s%-10s%-10s%-10s%-15s%-15s" % (
                    "GPU_name", "GPU_id", "Temp", "GPU-Util", "Memory-Usage", "Memory-Free"))
                for gpu in rets:
                    print("%-25s%-10s%-10s%-10s%-15s%-15s" % (
                        gpu[5], gpu[0], gpu[1], gpu[2] + "%", gpu[3] + "/" + gpu[4], int(gpu[4]) - int(gpu[3])))
            print("Ctrl+C 退出")
            print("================================================================================")
        except paramiko.ssh_exception.SSHException:
            conns = [
                SSHConnection(servers[name]['hostname'], port=servers[name]['port'], username=servers[name]['username']
                              , password=servers[name]['password']) for name in servers_name]
        except KeyboardInterrupt:
            break


def init_gpu():
    gpus = dict()
    gpu_count = 0
    for name in servers_name:
        gpus[name] = dict()
        for num in range(servers[name]['gpu_num']):
            gpus[name][num] = gpu_runner(name+"-gpu-"+str(num+1),
                                         gpu_count+1,
                                         {'hostname':servers[name]['hostname'],'port':servers[name]['port'],
                                          'username':servers[name]['username'],'password':servers[name]['password']},num)
            gpu_count += 1
    return gpus, gpu_count


def add_task(gpu_runners):
    name, gpu_id, free, dir, priority = 0,0,0,0,0
    while True:
        try:
            input_str = input("Please input server name, gpu_id, free, dir and priority split by space: \n").split(" ")
            name, gpu_id, free, dir, priority = input_str[0], int(input_str[1]),int(input_str[2]),input_str[3],int(input_str[4])
            if name not in servers_name:
                raise ValueError
            if gpu_id >= servers[name]['gpu_num']:
                raise ValueError
            break
        except KeyboardInterrupt:
            return False
        except Exception:
            print("Wrong gpu name or gpu_id")

    command = input("Please input command:\n")
    gpus_runners[name][gpu_id].append_task(Task(dir, command, free, priority, new_id()))
    print("OK")


def add_task_from_excel(gpu_runners):
    while True:
        excel_dir = 0
        try:
            input_str = input("Please input excel dir=:\n")
            excel_dir = input_str
            tasks = ExcelOp(excel_dir)
            break
        except KeyboardInterrupt:
            return False
        except FileNotFoundError:
            print("File Not Found")
        except Exception:
            print("Wrong")


    rn, cn = tasks.get_row_clo_num()
    for i in range(2, rn +1):
        if tasks.get_cell_value(i, 1) is not None:
            gpus_runners[tasks.get_cell_value(i, 4)][tasks.get_cell_value(i, 5)].\
                append_task(Task(tasks.get_cell_value(i, 2), tasks.get_cell_value(i, 3),
                                 tasks.get_cell_value(i, 6),tasks.get_cell_value(i, 7), new_id()))
        else:
            break
    print("OK")


def delete_task(gpu_runners):
    id = 0
    while True:
        try:
            id = int(input("Please input id: \n"))
            break
        except KeyboardInterrupt:
            return False
        except ValueError:
            print("Wrong gpu name or gpu_id")
    ok = False
    for name in servers_name:
        for gpu_id in list(gpu_runners[name].keys()):
            gpu = gpu_runners[name][gpu_id]
            ok = gpu.delete_task(id)
            if ok is True:
                print("OK")
                return True

    print("not ok")


def change_task(gpu_runners):
    id, free, dir, priority = 0, 0, 0, 0
    while True:
        try:
            input_str = input("Please input free, dir ,priority and id split by space:").split(
                " ")
            free, dir, priority, id = int(input_str[0]), input_str[1],int(input_str[2]), int(input_str[3])
            break
        except KeyboardInterrupt:
            return False
        except ValueError:
            print("Wrong gpu name or gpu_id")

    command = input("Please input command:\n")
    ok = False
    for name in servers_name:
        for gpu_id in list(gpu_runners[name].keys()):
            gpu = gpu_runners[name][gpu_id]
            if name == 'Mars' and gpu_id == 0:
                ok = gpu.change_task(id, Task(dir, command, free, priority, id))
                if ok is True:
                    print("OK")
                    return True
    print("not ok")


def show_running_task(gpu_runners):
    print("========================================================================================================"
          "===========================================================================")
    print("%-7s%-7s%-5s%-30s%-50s%-13s%-10s%-10s%-20s%-20s%-6s%-10s" % (
        "name", "GPU_id","id", "dir", "command", "state", "free", "priority", "start_time", "running_time", "ps_id", "used"))
    for name in servers_name:
        for gpu_id in list(gpu_runners[name].keys()):
            gpu = gpu_runners[name][gpu_id]
            for task in gpu.tasks:
                print("%-7s%-7d%-5d%-30s%-50s%-13s%-10s%-10d%-20s%-20s%-6d%-10s" % (
                    name, gpu.gpu_id, task.id, task.dir, task.command[:-1 if len(task.command) < 50 else 49],
                    task.state,task.free, task.priority, task.start_time, task.running_time,
                    task.ps_id, task.used))

    print("========================================================================================================"
          "===========================================================================")

def save_task(gpu_runners):
    while True:
        excel_dir = 0
        try:
            input_str = input("Please input excel dir:\n").split(" ")
            excel_dir = input_str[0]
            break
        except KeyboardInterrupt:
            return False
        except ValueError:
            print("Wrong")
    from openpyxl import  Workbook
    workbook = Workbook()
    # 默认sheet
    sheet = workbook.active
    sheet.title = "默认sheet"
    # 设置表头
    for i, key in enumerate(column):
        sheet.cell(1, i+1, key)
    row = 2
    for name in servers_name:
        for gpu_id in list(gpu_runners[name].keys()):
            gpu = gpu_runners[name][gpu_id]
            tasks = gpu.tasks
            for _, task in enumerate(tasks):
                sheet.cell(row, 1, task.id)
                sheet.cell(row, 2, task.dir)
                sheet.cell(row, 3, task.command)
                sheet.cell(row, 4, name)
                sheet.cell(row, 5, gpu_id)
                sheet.cell(row, 6, task.free)
                sheet.cell(row, 7, task.priority)
                sheet.cell(row, 8, task.state)
                sheet.cell(row, 9, task.start_time)
                sheet.cell(row, 10, task.running_time)
                sheet.cell(row, 11, task.ps_id)
                sheet.cell(row, 12, task.used)
                row += 1
    workbook.save(excel_dir)
    print("Ok")

def task_manager_menu():
    print("1- Add New Task")
    print("2- Add New Task from xlsx")
    print("3- Delete Task")
    print("4- Change")
    print("5- Show All Tasks")
    print("6- Save To Excel")
    print("7- exit")
    ia = input("Enter Value 1-7: ")
    try:
        ia = int(ia)
        if ia < 1 or ia > 7:
            print("Error Input,please Enter Value 1-7:", end='')
        else:
            return ia
    except ValueError:
        print("Error Input,please Enter Value 1-7:", end='')

    ia = input()
    while True:
        try:
            ia = int(ia)
            if ia < 1 or ia > 7:
                ia = input("Error Input,please Enter Value 1-7: ")
            else:
                break
        except ValueError:
            ia = input("Error Input,please Enter Value 1-7: ")
    return ia


def task_manager(gpu_runners):
    while True:
        ia = task_manager_menu()
        if ia == 1:
            add_task(gpu_runners)
        elif ia == 2:
            add_task_from_excel(gpu_runners)
        elif ia == 3:
            delete_task(gpu_runners)
        elif ia == 4:
            change_task(gpu_runners)
        elif ia == 5:
            show_running_task(gpu_runners)
        elif ia == 6:
            save_task(gpu_runners)
        elif ia == 7:
            break


def new_id():
    global task_id
    task_id = task_id + 1
    return task_id

def runner_menu():
    print("1- Start")
    print("2- Pause")
    print("3- Resume")
    print("4- Stop")
    print("5- exit")
    ia = input("Enter Value 1-5: ")
    try:
        ia = int(ia)
        if ia < 1 or ia > 5:
            print("Error Input,please Enter Value 1-5:", end='')
        else:
            return ia
    except ValueError:
        print("Error Input,please Enter Value 1-5:", end='')

    ia = input()
    while True:
        try:
            ia = int(ia)
            if ia < 1 or ia > 5:
                ia = input("Error Input,please Enter Value 1-5: ")
            else:
                break
        except ValueError:
            ia = input("Error Input,please Enter Value 1-5: ")
    return ia

def runner(gpu_runners):
    while True:
        ia = runner_menu()

        if ia == 1:
            for name in servers_name:
                for gpu_id in list(gpu_runners[name].keys()):
                    gpu_runners[name][gpu_id].start()

        elif ia == 2:
            for name in servers_name:
                for gpu_id in list(gpu_runners[name].keys()):
                    gpu_runners[name][gpu_id].pause()
        elif ia == 3:
            for name in servers_name:
                for gpu_id in list(gpu_runners[name].keys()):
                    gpu_runners[name][gpu_id].resume()
        elif ia == 4:
            for name in servers_name:
                for gpu_id in list(gpu_runners[name].keys()):
                    gpu_runners[name][gpu_id].stop()

        elif ia == 5:
            break



if __name__ == "__main__":
    welcome()
    print("init—system.......")
    gpus_runners, gpu_nums = init_gpu()
    while True:
        ia = print_menu()
        if ia == 1:
            monitor()
        elif ia == 2:
            print("===========================================================")
            task_manager(gpus_runners)
        elif ia == 3:
            print("===========================================================")
            runner(gpus_runners)
        elif ia == 4:
            for name in servers_name:
                for gpu_id in list(gpus_runners[name].keys()):
                    gpus_runners[name][gpu_id].stop()
            exit()


        # ia = print_menu()

